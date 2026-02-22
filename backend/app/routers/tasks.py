import io
import os
import re
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, case
from typing import Optional, List
from datetime import date
from pydantic import BaseModel
from app.database import get_db
from app.models.user import User, TaskFavorite, UserNotificationPreference
from app.models.task import Task, TaskLog, Notification, Category, Tag, StatusEnum, PriorityEnum, task_tags, Mention
from app.utils.auth import get_current_user
from app.utils.email import send_task_assigned, send_status_changed
from app.config import get_settings

_settings = get_settings()
router = APIRouter(prefix="/api/tasks", tags=["tasks"])

# ── 검색 키워드 → 필터 매핑 ────────────────────────────────
SEARCH_STATUS_KEYWORDS: dict[str, StatusEnum] = {
    "대기": StatusEnum.pending,
    "진행중": StatusEnum.in_progress,
    "진행": StatusEnum.in_progress,
    "검토": StatusEnum.review,
    "검토요청": StatusEnum.review,
    "검토중": StatusEnum.review,
    "완료": StatusEnum.approved,
    "승인": StatusEnum.approved,
    "반려": StatusEnum.rejected,
    "반려됨": StatusEnum.rejected,
}

SEARCH_PRIORITY_KEYWORDS: dict[str, PriorityEnum] = {
    "긴급": PriorityEnum.urgent,
    "급함": PriorityEnum.urgent,
    "급한": PriorityEnum.urgent,
    "높음": PriorityEnum.high,
    "보통": PriorityEnum.normal,
    "낮음": PriorityEnum.low,
}


# ── Schemas ──────────────────────────────────────────────
class MaterialProvider(BaseModel):
    assignee_id: int
    title: str
    due_date: Optional[date] = None


class TaskCreate(BaseModel):
    title: str
    content: Optional[str] = None
    assignee_id: int
    category_id: Optional[int] = None
    priority: PriorityEnum = PriorityEnum.normal
    estimated_hours: Optional[float] = None
    due_date: Optional[date] = None
    tag_names: List[str] = []
    parent_task_id: Optional[int] = None
    is_subtask: bool = False
    material_providers: List[MaterialProvider] = []


class TaskUpdate(BaseModel):
    title: Optional[str] = None
    content: Optional[str] = None
    assignee_id: Optional[int] = None
    category_id: Optional[int] = None
    priority: Optional[PriorityEnum] = None
    estimated_hours: Optional[float] = None
    due_date: Optional[date] = None
    progress: Optional[int] = None
    tag_names: Optional[List[str]] = None


class CommentRequest(BaseModel):
    comment: str


class CommentUpdate(BaseModel):
    comment: str


class StatusChange(BaseModel):
    status: StatusEnum
    comment: Optional[str] = None


class BulkStatusChange(BaseModel):
    task_ids: List[int]
    status: StatusEnum
    comment: Optional[str] = None


# ── Helpers ──────────────────────────────────────────────
def task_to_dict(t: Task, current_user_id: Optional[int] = None) -> dict:
    return {
        "id": t.id,
        "title": t.title,
        "content": t.content,
        "assigner": {"id": t.assigner.id, "name": t.assigner.name} if t.assigner else None,
        "assignee": {"id": t.assignee.id, "name": t.assignee.name} if t.assignee else None,
        "category": {"id": t.category.id, "name": t.category.name, "color": t.category.color} if t.category else None,
        "priority": t.priority,
        "status": t.status,
        "progress": t.progress,
        "estimated_hours": float(t.estimated_hours) if t.estimated_hours else None,
        "due_date": t.due_date.isoformat() if t.due_date else None,
        "is_subtask": t.is_subtask,
        "parent_task_id": t.parent_task_id,
        "tags": [{"id": tag.id, "name": tag.name} for tag in t.tags],
        "attachment_count": len(t.attachments),
        "subtask_count": len(t.subtasks),
        "subtasks_done": sum(1 for s in t.subtasks if s.status == StatusEnum.approved),
        "is_favorite": any(f.user_id == current_user_id for f in t.favorites) if current_user_id else False,
        "created_at": t.created_at.isoformat() if t.created_at else None,
        "updated_at": t.updated_at.isoformat() if t.updated_at else None,
    }


def _get_or_create_tags(db: Session, names: List[str]) -> List[Tag]:
    tags = []
    for name in names:
        name = name.strip()
        if not name:
            continue
        tag = db.query(Tag).filter(Tag.name == name).first()
        if not tag:
            tag = Tag(name=name)
            db.add(tag)
            db.flush()
        tags.append(tag)
    return tags


def _add_notification(db: Session, user_id: int, task_id: int, ntype: str, message: str):
    notif = Notification(user_id=user_id, task_id=task_id, type=ntype, message=message)
    db.add(notif)


def _check_notify_pref(db: Session, user_id: int, pref_field: str) -> bool:
    """사용자의 알림 설정을 확인. 설정 없으면 기본 True."""
    pref = db.query(UserNotificationPreference).filter(
        UserNotificationPreference.user_id == user_id
    ).first()
    if not pref:
        return True
    return getattr(pref, pref_field, True)


def _check_task_access(task: Task, current_user: User) -> bool:
    return (
        current_user.id == task.assigner_id
        or current_user.id == task.assignee_id
        or current_user.is_admin
    )


def _parse_mentions(comment: str, db: Session) -> List[User]:
    """댓글에서 @이름 패턴을 파싱해 실제 사용자 목록 반환"""
    mention_names = set(re.findall(r"@(\S+)", comment))
    mentioned_users = []
    for name in mention_names:
        user = db.query(User).filter(User.name == name, User.is_active == True).first()
        if user:
            mentioned_users.append(user)
    return mentioned_users


# ── Endpoints ─────────────────────────────────────────────
@router.post("/", status_code=201)
async def create_task(req: TaskCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    assignee = db.query(User).filter(User.id == req.assignee_id).first()
    if not assignee:
        raise HTTPException(status_code=404, detail="담당자를 찾을 수 없습니다.")

    task = Task(
        title=req.title,
        content=req.content,
        assigner_id=current_user.id,
        assignee_id=req.assignee_id,
        category_id=req.category_id,
        priority=req.priority,
        estimated_hours=req.estimated_hours,
        due_date=req.due_date,
        parent_task_id=req.parent_task_id,
        is_subtask=req.is_subtask,
    )
    task.tags = _get_or_create_tags(db, req.tag_names)
    db.add(task)
    db.flush()

    db.add(TaskLog(task_id=task.id, user_id=current_user.id, action="created", new_value="pending"))
    if _check_notify_pref(db, assignee.id, "notify_assigned"):
        _add_notification(db, assignee.id, task.id, "assigned", f"새 업무가 배정되었습니다: {task.title}")

    material_assignees = []
    for mp in req.material_providers:
        provider = db.query(User).filter(User.id == mp.assignee_id).first()
        if not provider:
            continue
        subtask = Task(
            title=mp.title,
            content=f"[자료제공] {task.title} 업무의 자료 제공 요청",
            assigner_id=current_user.id,
            assignee_id=mp.assignee_id,
            priority=req.priority,
            due_date=mp.due_date or req.due_date,
            parent_task_id=task.id,
            is_subtask=True,
        )
        db.add(subtask)
        db.flush()
        db.add(TaskLog(task_id=subtask.id, user_id=current_user.id, action="created", new_value="pending"))
        if _check_notify_pref(db, mp.assignee_id, "notify_assigned"):
            _add_notification(db, mp.assignee_id, subtask.id, "assigned", f"자료 제공 요청: {mp.title}")
        material_assignees.append((provider, subtask))

    db.commit()

    try:
        if _check_notify_pref(db, assignee.id, "email_assigned"):
            await send_task_assigned(
                to=assignee.email,
                assignee_name=assignee.name,
                task_title=task.title,
                assigner_name=current_user.name,
                due_date=str(req.due_date) if req.due_date else "미정",
            )
    except Exception:
        pass

    for provider, subtask in material_assignees:
        try:
            if _check_notify_pref(db, provider.id, "email_assigned"):
                await send_task_assigned(
                    to=provider.email,
                    assignee_name=provider.name,
                    task_title=subtask.title,
                    assigner_name=current_user.name,
                    due_date=str(subtask.due_date) if subtask.due_date else "미정",
                )
        except Exception:
            pass

    return task_to_dict(
        db.query(Task).options(
            joinedload(Task.assigner), joinedload(Task.assignee),
            joinedload(Task.category), joinedload(Task.tags),
            joinedload(Task.attachments), joinedload(Task.subtasks),
            joinedload(Task.favorites),
        ).filter(Task.id == task.id).first(),
        current_user_id=current_user.id,
    )


@router.get("/export")
def export_tasks(
    section: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    due_date_from: Optional[date] = Query(None),
    due_date_to: Optional[date] = Query(None),
    fmt: str = Query("xlsx", regex="^(xlsx|csv)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """업무 목록을 Excel(.xlsx) 또는 CSV로 내보내기"""
    STATUS_LABELS = {
        "pending": "대기", "in_progress": "진행중", "review": "검토요청",
        "approved": "완료", "rejected": "반려",
    }
    PRIORITY_LABELS = {
        "urgent": "긴급", "high": "높음", "normal": "보통", "low": "낮음",
    }

    query = db.query(Task).options(
        joinedload(Task.assigner), joinedload(Task.assignee),
        joinedload(Task.category), joinedload(Task.tags),
    )

    if section == "assigned_to_me":
        query = query.filter(Task.assignee_id == current_user.id, Task.is_subtask == False)
    elif section == "assigned_by_me":
        query = query.filter(Task.assigner_id == current_user.id, Task.is_subtask == False)
    elif section == "subtasks_to_me":
        query = query.filter(Task.assignee_id == current_user.id, Task.is_subtask == True)
    else:
        # 관리자: 전체 / 일반 사용자: 본인 관련
        if not current_user.is_admin:
            query = query.filter(
                or_(Task.assigner_id == current_user.id, Task.assignee_id == current_user.id)
            )

    if status:
        query = query.filter(Task.status == status)
    if priority:
        query = query.filter(Task.priority == priority)
    if search:
        query = query.filter(or_(Task.title.contains(search), Task.content.contains(search)))
    if due_date_from:
        query = query.filter(Task.due_date >= due_date_from)
    if due_date_to:
        query = query.filter(Task.due_date <= due_date_to)

    tasks = query.order_by(Task.created_at.desc()).limit(5000).all()

    headers_map = [
        ("번호", lambda t: t.id),
        ("제목", lambda t: t.title),
        ("지시자", lambda t: t.assigner.name if t.assigner else ""),
        ("담당자", lambda t: t.assignee.name if t.assignee else ""),
        ("상태", lambda t: STATUS_LABELS.get(t.status.value if t.status else "", "")),
        ("우선순위", lambda t: PRIORITY_LABELS.get(t.priority.value if t.priority else "", "")),
        ("카테고리", lambda t: t.category.name if t.category else ""),
        ("태그", lambda t: ", ".join(tag.name for tag in t.tags)),
        ("진행도(%)", lambda t: t.progress or 0),
        ("마감일", lambda t: t.due_date.isoformat() if t.due_date else ""),
        ("생성일", lambda t: t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else ""),
    ]

    if fmt == "csv":
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([h for h, _ in headers_map])
        for t in tasks:
            writer.writerow([fn(t) for _, fn in headers_map])
        content = "\ufeff" + output.getvalue()  # BOM for Excel UTF-8
        return StreamingResponse(
            io.BytesIO(content.encode("utf-8")),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=tasks.csv"},
        )
    else:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "업무 목록"

        # 헤더
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col, (header, _) in enumerate(headers_map, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 데이터
        for row, t in enumerate(tasks, 2):
            for col, (_, fn) in enumerate(headers_map, 1):
                ws.cell(row=row, column=col, value=fn(t))

        # 컬럼 너비 자동 조정
        col_widths = [8, 40, 12, 12, 10, 10, 12, 20, 10, 12, 18]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=tasks.xlsx"},
        )


@router.get("/favorites")
def list_favorites(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """내 즐겨찾기 업무 목록"""
    favs = (
        db.query(Task)
        .join(TaskFavorite, TaskFavorite.task_id == Task.id)
        .options(
            joinedload(Task.assigner), joinedload(Task.assignee),
            joinedload(Task.category), joinedload(Task.tags),
            joinedload(Task.attachments), joinedload(Task.subtasks),
            joinedload(Task.favorites),
        )
        .filter(TaskFavorite.user_id == current_user.id)
        .order_by(TaskFavorite.created_at.desc())
        .all()
    )
    return [task_to_dict(t, current_user.id) for t in favs]


@router.get("/dashboard")
def dashboard_summary(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from datetime import timedelta
    today = date.today()
    in_3_days = today + date.fromordinal(today.toordinal() + 3) - today  # timedelta(days=3)
    in_3_days = today.__class__.fromordinal(today.toordinal() + 3)

    my_tasks = db.query(Task).filter(Task.assignee_id == current_user.id, Task.is_subtask == False)
    total = my_tasks.count()
    urgent = my_tasks.filter(Task.priority == PriorityEnum.urgent, Task.status != StatusEnum.approved).count()
    due_soon = my_tasks.filter(
        Task.due_date <= in_3_days,
        Task.due_date >= today,
        Task.status.notin_([StatusEnum.approved]),
    ).count()
    overdue = my_tasks.filter(
        Task.due_date < today,
        Task.status.notin_([StatusEnum.approved]),
    ).count()
    rejected = my_tasks.filter(Task.status == StatusEnum.rejected).count()
    status_breakdown = {s.value: my_tasks.filter(Task.status == s).count() for s in StatusEnum}

    return {
        "total": total,
        "urgent": urgent,
        "due_soon": due_soon,
        "overdue": overdue,
        "rejected": rejected,
        "status_breakdown": status_breakdown,
    }


@router.get("/")
def list_tasks(
    section: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    assignee_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    due_date_from: Optional[date] = Query(None),
    due_date_to: Optional[date] = Query(None),
    sort_by: Optional[str] = Query(None),
    sort_dir: str = Query("desc"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    favorites_only: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    text_terms: List[str] = []
    implied_status: Optional[StatusEnum] = None
    implied_priority: Optional[PriorityEnum] = None

    if search:
        for word in search.strip().split():
            if word in SEARCH_STATUS_KEYWORDS and status is None and implied_status is None:
                implied_status = SEARCH_STATUS_KEYWORDS[word]
            elif word in SEARCH_PRIORITY_KEYWORDS and priority is None and implied_priority is None:
                implied_priority = SEARCH_PRIORITY_KEYWORDS[word]
            else:
                text_terms.append(word)

    def apply_filters(q):
        if favorites_only:
            q = q.join(TaskFavorite, (TaskFavorite.task_id == Task.id) & (TaskFavorite.user_id == current_user.id))
        elif section == "assigned_to_me":
            q = q.filter(Task.assignee_id == current_user.id, Task.is_subtask == False)
        elif section == "assigned_by_me":
            q = q.filter(Task.assigner_id == current_user.id, Task.is_subtask == False)
        elif section == "subtasks_to_me":
            q = q.filter(Task.assignee_id == current_user.id, Task.is_subtask == True)

        if status:
            q = q.filter(Task.status == status)
        elif implied_status:
            q = q.filter(Task.status == implied_status)
        if priority:
            q = q.filter(Task.priority == priority)
        elif implied_priority:
            q = q.filter(Task.priority == implied_priority)
        if assignee_id:
            q = q.filter(Task.assignee_id == assignee_id)

        for term in text_terms:
            user_ids_matching = db.query(User.id).filter(User.name.contains(term))
            tag_ids_matching = db.query(Tag.id).filter(Tag.name.contains(term))
            cat_ids_matching = db.query(Category.id).filter(Category.name.contains(term))
            task_ids_with_tag = db.query(task_tags.c.task_id).filter(
                task_tags.c.tag_id.in_(tag_ids_matching)
            )
            q = q.filter(or_(
                Task.title.contains(term),
                Task.content.contains(term),
                Task.assignee_id.in_(user_ids_matching),
                Task.assigner_id.in_(user_ids_matching),
                Task.category_id.in_(cat_ids_matching),
                Task.id.in_(task_ids_with_tag),
            ))

        if due_date_from:
            q = q.filter(Task.due_date >= due_date_from)
        if due_date_to:
            q = q.filter(Task.due_date <= due_date_to)
        return q

    PRIORITY_ORDER = case(
        (Task.priority == PriorityEnum.urgent, 1),
        (Task.priority == PriorityEnum.high, 2),
        (Task.priority == PriorityEnum.normal, 3),
        (Task.priority == PriorityEnum.low, 4),
        else_=5,
    )
    STATUS_ORDER = case(
        (Task.status == StatusEnum.rejected, 1),
        (Task.status == StatusEnum.pending, 2),
        (Task.status == StatusEnum.in_progress, 3),
        (Task.status == StatusEnum.review, 4),
        (Task.status == StatusEnum.approved, 5),
        else_=6,
    )
    asc_flag = sort_dir == "asc"
    DUE_DATE_NULL_FLAG = case((Task.due_date.is_(None), 1), else_=0)
    sort_map = {
        "due_date": (DUE_DATE_NULL_FLAG.asc(), Task.due_date.asc()) if asc_flag else (DUE_DATE_NULL_FLAG.asc(), Task.due_date.desc()),
        "priority": PRIORITY_ORDER.asc() if asc_flag else PRIORITY_ORDER.desc(),
        "status": STATUS_ORDER.asc() if asc_flag else STATUS_ORDER.desc(),
        "title": Task.title.asc() if asc_flag else Task.title.desc(),
        "created_at": Task.created_at.asc() if asc_flag else Task.created_at.desc(),
    }
    order_expr = sort_map.get(sort_by, Task.created_at.desc())
    order_clause = list(order_expr) if isinstance(order_expr, tuple) else [order_expr]

    total = apply_filters(db.query(Task)).count()
    items = apply_filters(
        db.query(Task).options(
            joinedload(Task.assigner), joinedload(Task.assignee),
            joinedload(Task.category), joinedload(Task.tags),
            joinedload(Task.attachments), joinedload(Task.subtasks),
            joinedload(Task.favorites),
        )
    ).order_by(*order_clause).offset((page - 1) * page_size).limit(page_size).all()

    return {"items": [task_to_dict(t, current_user.id) for t in items], "total": total, "page": page, "page_size": page_size}


@router.post("/bulk-status")
async def bulk_change_status(
    req: BulkStatusChange,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if req.status == StatusEnum.rejected and not req.comment:
        raise HTTPException(status_code=400, detail="반려 시 코멘트는 필수입니다.")
    if not req.task_ids:
        raise HTTPException(status_code=400, detail="업무를 선택해주세요.")

    updated = 0
    for task_id in req.task_ids:
        task = db.query(Task).options(
            joinedload(Task.assigner), joinedload(Task.assignee)
        ).filter(Task.id == task_id).first()
        if not task or not _check_task_access(task, current_user):
            continue

        old_status = task.status
        task.status = req.status
        db.add(TaskLog(
            task_id=task.id,
            user_id=current_user.id,
            action="status_changed",
            old_value=old_status,
            new_value=req.status,
            comment=req.comment,
        ))
        notify_user_id = task.assigner_id if current_user.id == task.assignee_id else task.assignee_id
        if _check_notify_pref(db, notify_user_id, "notify_status_changed"):
            _add_notification(db, notify_user_id, task.id, "status_changed",
                              f"업무 상태 변경: {task.title} → {req.status}")
        updated += 1

    db.commit()
    return {"message": f"{updated}건의 업무 상태가 변경되었습니다.", "updated": updated}


@router.get("/{task_id}")
def get_task(task_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    task = db.query(Task).options(
        joinedload(Task.assigner), joinedload(Task.assignee),
        joinedload(Task.category), joinedload(Task.tags),
        joinedload(Task.attachments), joinedload(Task.subtasks),
        joinedload(Task.logs), joinedload(Task.favorites),
    ).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")
    # 접근 제어: 지시자, 담당자, 관리자만 조회 가능
    if not _check_task_access(task, current_user):
        raise HTTPException(status_code=403, detail="이 업무에 접근할 권한이 없습니다.")
    return task_to_dict(task, current_user.id)


@router.post("/{task_id}/favorite")
def toggle_favorite(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")
    if not _check_task_access(task, current_user):
        raise HTTPException(status_code=403, detail="이 업무에 접근할 권한이 없습니다.")

    fav = db.query(TaskFavorite).filter(
        TaskFavorite.user_id == current_user.id,
        TaskFavorite.task_id == task_id,
    ).first()

    if fav:
        db.delete(fav)
        db.commit()
        return {"is_favorite": False, "message": "즐겨찾기가 해제되었습니다."}
    else:
        db.add(TaskFavorite(user_id=current_user.id, task_id=task_id))
        db.commit()
        return {"is_favorite": True, "message": "즐겨찾기에 추가되었습니다."}


@router.patch("/{task_id}")
def update_task(
    task_id: int,
    req: TaskUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")
    if not _check_task_access(task, current_user):
        raise HTTPException(status_code=403, detail="업무를 수정할 권한이 없습니다.")

    if req.title is not None:
        task.title = req.title
    if req.content is not None:
        task.content = req.content
    if req.category_id is not None:
        task.category_id = req.category_id
    if req.priority is not None:
        task.priority = req.priority
    if req.estimated_hours is not None:
        task.estimated_hours = req.estimated_hours
    if req.due_date is not None:
        task.due_date = req.due_date
    if req.progress is not None:
        task.progress = max(0, min(100, req.progress))
    if req.tag_names is not None:
        task.tags = _get_or_create_tags(db, req.tag_names)

    reassigned = False
    if req.assignee_id is not None and current_user.id == task.assigner_id:
        new_assignee = db.query(User).filter(User.id == req.assignee_id, User.is_active == True).first()
        if not new_assignee:
            raise HTTPException(status_code=404, detail="담당자를 찾을 수 없습니다.")
        old_assignee_id = task.assignee_id
        task.assignee_id = req.assignee_id
        db.add(TaskLog(task_id=task.id, user_id=current_user.id, action="reassigned",
                       old_value=str(old_assignee_id), new_value=str(req.assignee_id)))
        if _check_notify_pref(db, req.assignee_id, "notify_reassigned"):
            _add_notification(db, req.assignee_id, task.id, "reassigned", f"업무가 재배정되었습니다: {task.title}")
        reassigned = True

    other_fields_changed = any(v is not None for v in [
        req.title, req.content, req.category_id, req.priority,
        req.estimated_hours, req.due_date, req.progress, req.tag_names,
    ])
    if other_fields_changed or not reassigned:
        db.add(TaskLog(task_id=task.id, user_id=current_user.id, action="updated"))
    db.commit()
    return {"message": "업무가 수정되었습니다."}


@router.post("/{task_id}/status")
async def change_status(
    task_id: int,
    req: StatusChange,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    task = db.query(Task).options(
        joinedload(Task.assigner), joinedload(Task.assignee)
    ).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")
    if not _check_task_access(task, current_user):
        raise HTTPException(status_code=403, detail="상태를 변경할 권한이 없습니다.")
    if req.status == StatusEnum.rejected and not req.comment:
        raise HTTPException(status_code=400, detail="반려 시 코멘트는 필수입니다.")

    old_status = task.status
    task.status = req.status
    db.add(TaskLog(
        task_id=task.id,
        user_id=current_user.id,
        action="status_changed",
        old_value=old_status,
        new_value=req.status,
        comment=req.comment,
    ))

    notify_user_id = task.assigner_id if current_user.id == task.assignee_id else task.assignee_id
    if _check_notify_pref(db, notify_user_id, "notify_status_changed"):
        _add_notification(db, notify_user_id, task.id, "status_changed",
                          f"업무 상태 변경: {task.title} → {req.status}")
    db.commit()

    notify_user = db.query(User).filter(User.id == notify_user_id).first()
    try:
        if notify_user and _check_notify_pref(db, notify_user_id, "email_status_changed"):
            await send_status_changed(
                to=notify_user.email,
                user_name=notify_user.name,
                task_title=task.title,
                old_status=old_status,
                new_status=req.status,
            )
    except Exception:
        pass

    return {"message": "상태가 변경되었습니다."}


@router.get("/{task_id}/logs")
def get_logs(task_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")
    if not _check_task_access(task, current_user):
        raise HTTPException(status_code=403, detail="이 업무에 접근할 권한이 없습니다.")

    logs = (
        db.query(TaskLog)
        .options(joinedload(TaskLog.user), joinedload(TaskLog.mentions).joinedload(Mention.mentioned_user))
        .filter(TaskLog.task_id == task_id)
        .order_by(TaskLog.created_at.asc())
        .all()
    )
    return [
        {
            "id": l.id,
            "user": {"id": l.user.id, "name": l.user.name} if l.user else None,
            "action": l.action,
            "old_value": l.old_value,
            "new_value": l.new_value,
            "comment": l.comment,
            "mentions": [
                {"user_id": m.mentioned_user_id, "name": m.mentioned_user.name}
                for m in l.mentions
            ],
            "created_at": l.created_at.isoformat(),
        }
        for l in logs
    ]


@router.post("/{task_id}/comment")
def add_comment(
    task_id: int,
    req: CommentRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")
    if not _check_task_access(task, current_user):
        raise HTTPException(status_code=403, detail="댓글을 작성할 권한이 없습니다.")
    if not req.comment.strip():
        raise HTTPException(status_code=400, detail="댓글 내용을 입력해주세요.")

    log = TaskLog(
        task_id=task.id,
        user_id=current_user.id,
        action="comment",
        comment=req.comment.strip(),
    )
    db.add(log)
    db.flush()

    # @멘션 파싱 후 알림 생성
    mentioned_users = _parse_mentions(req.comment, db)
    for mentioned_user in mentioned_users:
        if mentioned_user.id == current_user.id:
            continue
        db.add(Mention(log_id=log.id, mentioned_user_id=mentioned_user.id))
        if _check_notify_pref(db, mentioned_user.id, "notify_mentioned"):
            _add_notification(
                db, mentioned_user.id, task_id, "mentioned",
                f"{current_user.name}님이 댓글에서 멘션했습니다: {task.title}",
            )

    # 댓글 알림: 상대방에게 알림 (멘션 외)
    other_party_id = task.assigner_id if current_user.id == task.assignee_id else task.assignee_id
    mentioned_ids = {u.id for u in mentioned_users}
    if other_party_id not in mentioned_ids and other_party_id != current_user.id:
        if _check_notify_pref(db, other_party_id, "notify_commented"):
            _add_notification(
                db, other_party_id, task_id, "commented",
                f"{current_user.name}님이 댓글을 작성했습니다: {task.title}",
            )

    db.commit()
    return {"message": "댓글이 등록되었습니다.", "log_id": log.id}


@router.post("/{task_id}/clone", status_code=201)
async def clone_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    original = db.query(Task).options(joinedload(Task.tags)).filter(Task.id == task_id).first()
    if not original:
        raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")
    if not _check_task_access(original, current_user):
        raise HTTPException(status_code=403, detail="업무를 복사할 권한이 없습니다.")

    cloned = Task(
        title=f"[복사] {original.title}",
        content=original.content,
        assigner_id=current_user.id,
        assignee_id=original.assignee_id,
        category_id=original.category_id,
        priority=original.priority,
        estimated_hours=original.estimated_hours,
        due_date=original.due_date,
        is_subtask=original.is_subtask,
        parent_task_id=original.parent_task_id,
    )
    cloned.tags = list(original.tags)
    db.add(cloned)
    db.flush()
    db.add(TaskLog(task_id=cloned.id, user_id=current_user.id, action="created", new_value="pending"))
    if _check_notify_pref(db, original.assignee_id, "notify_assigned"):
        _add_notification(db, original.assignee_id, cloned.id, "assigned", f"업무가 복사 배정되었습니다: {cloned.title}")
    db.commit()
    return {"message": "업무가 복사되었습니다.", "task_id": cloned.id}


@router.delete("/{task_id}", status_code=204)
def delete_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    task = db.query(Task).options(
        joinedload(Task.subtasks).joinedload(Task.attachments),
        joinedload(Task.attachments),
    ).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")
    if current_user.id != task.assigner_id and not current_user.is_admin:
        raise HTTPException(status_code=403, detail="업무를 삭제할 권한이 없습니다.")

    all_attachments = list(task.attachments)
    for subtask in task.subtasks:
        all_attachments.extend(subtask.attachments)
    for att in all_attachments:
        fpath = os.path.join(_settings.file_storage_path, att.stored_name)
        if os.path.exists(fpath):
            try:
                os.remove(fpath)
            except OSError:
                pass

    for subtask in list(task.subtasks):
        db.delete(subtask)
    db.flush()
    db.delete(task)
    db.commit()


@router.patch("/{task_id}/comments/{log_id}")
def update_comment(
    task_id: int,
    log_id: int,
    req: CommentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    log = db.query(TaskLog).filter(
        TaskLog.id == log_id,
        TaskLog.task_id == task_id,
        TaskLog.action == "comment",
    ).first()
    if not log:
        raise HTTPException(status_code=404, detail="댓글을 찾을 수 없습니다.")
    if log.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="본인 댓글만 수정할 수 있습니다.")
    if not req.comment.strip():
        raise HTTPException(status_code=400, detail="댓글 내용을 입력해주세요.")
    log.comment = req.comment.strip()
    db.commit()
    return {"message": "댓글이 수정되었습니다."}


@router.delete("/{task_id}/comments/{log_id}", status_code=204)
def delete_comment(
    task_id: int,
    log_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    log = db.query(TaskLog).filter(
        TaskLog.id == log_id,
        TaskLog.task_id == task_id,
        TaskLog.action == "comment",
    ).first()
    if not log:
        raise HTTPException(status_code=404, detail="댓글을 찾을 수 없습니다.")
    if log.user_id != current_user.id and not current_user.is_admin:
        raise HTTPException(status_code=403, detail="본인 댓글만 삭제할 수 있습니다.")
    db.delete(log)
    db.commit()
